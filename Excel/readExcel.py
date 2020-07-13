import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

# Getting sheets form the workbook

# print(wb.sheetnames)
#
# for sheet in wb:
#     print(sheet.title)
#
# mySheet = wb.create_sheet("mySheet")
# print(wb.sheetnames)

# sheets = wb.get_sheet_by_name("Sheet3")
# sheet4 = wb["mySheet"]

# Getting cell from the sheets
ws = wb.active
# print(ws)
# print(ws['A2'])
# print(ws['A2'].value)

# c = ws['B1']
# print('Row {}, Column {} is {}\n'.format(c.row, c.column, c.value))
# print('Cell {},is {}\n'.format(c.coordinate, c.value))
# print(ws.cell(row=1, column=2).value)
# for i in range(1, 17, 1):
#     print(i,ws.cell(row=i, column=2).value)

# Getting rows and columns from sheets
colC = ws['B']
row6 = ws[6]
col_range = ws['A:B']
row_range = ws[1:4]
# print(colC[2].value)
for col in col_range:
    for cell in col:
        print(cell.value)

for row in row_range:
    for cell in row:
        print(cell.value)
